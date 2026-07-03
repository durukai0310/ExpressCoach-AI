#!/usr/bin/env python3
"""Generate 5 charts for W5 analysis, export PNG + embed in Excel."""

import os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
from matplotlib.patches import FancyBboxPatch
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from io import BytesIO

OUT = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# Global style
# ============================================================
plt.rcParams.update({
    'font.family': 'sans-serif',
    'font.sans-serif': ['Microsoft YaHei', 'SimHei', 'DejaVu Sans', 'Arial'],
    'font.size': 11,
    'axes.titlesize': 15,
    'axes.titleweight': 'bold',
    'axes.labelsize': 12,
    'axes.spines.top': False,
    'axes.spines.right': False,
    'figure.dpi': 150,
    'savefig.dpi': 150,
    'savefig.bbox': 'tight',
    'savefig.pad_inches': 0.2,
})

# Brand colors
EC_BLUE   = '#2563EB'  # ExpressCoach
GPT_ORANGE = '#F97316'  # GPT-4o
CLAUDE_GREEN = '#10B981'  # Claude
DS_BLACK  = '#374151'  # DeepSeek

GREEN_PASS = '#22C55E'
YELLOW_PARTIAL = '#EAB308'
RED_FAIL  = '#EF4444'

PALETTE_5 = ['#2563EB', '#F97316', '#10B981', '#374151', '#8B5CF6']


def save_png(fig, name):
    path = os.path.join(OUT, name)
    fig.savefig(path, facecolor='white', edgecolor='none')
    print(f'  [OK] Saved: {path}')
    return path


def embed_to_excel(image_paths, xlsx_path):
    """Embed PNGs into a single Excel sheet, one per row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'W5 Figures'

    # Title row
    ws.merge_cells('A1:F1')
    ws['A1'] = 'W5 Analysis Figures'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=16, color='2563EB')

    for i, (title, path) in enumerate(image_paths):
        row = 3 + i * 18  # leave room for each image
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=13, color='374151')

        img = XLImage(path)
        # Scale to fit ~700px wide
        img.width = 700
        img.height = 480
        ws.add_image(img, f'A{row + 1}')

    # Column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(xlsx_path)
    print(f'\n[OK] Excel saved: {xlsx_path}')


# ============================================================
# FIGURE 1 — Radar Chart: ExpressCoach vs GPT-4o vs Claude vs DeepSeek
# ============================================================
def fig1_radar():
    categories = ['Accuracy', 'Empathy', 'Speed', 'Personalization',
                  'Context\nMemory', 'Safety', 'Multilinguality']
    N = len(categories)

    # Data (0-100 scale)
    ec  = [88, 92, 78, 95, 85, 90, 80]
    gpt = [82, 70, 95, 72, 88, 75, 92]
    cld = [85, 88, 70, 80, 92, 95, 78]
    ds  = [78, 65, 90, 68, 80, 72, 85]

    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    angles += angles[:1]  # close the loop

    # Close each dataset
    def close(vals):
        return vals + vals[:1]

    fig, ax = plt.subplots(figsize=(7.5, 7.5), subplot_kw={'projection': 'polar'})
    fig.patch.set_facecolor('white')

    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)

    ax.plot(angles, close(ec),  'o-', color=EC_BLUE,   linewidth=2.2, markersize=6, label='ExpressCoach')
    ax.plot(angles, close(gpt), 's-', color=GPT_ORANGE, linewidth=2.2, markersize=6, label='GPT-4o')
    ax.plot(angles, close(cld), 'D-', color=CLAUDE_GREEN, linewidth=2.2, markersize=6, label='Claude')
    ax.plot(angles, close(ds),  'v-', color=DS_BLACK,  linewidth=2.2, markersize=6, label='DeepSeek')

    ax.fill(angles, close(ec),  EC_BLUE,   alpha=0.08)
    ax.fill(angles, close(gpt), GPT_ORANGE, alpha=0.08)
    ax.fill(angles, close(cld), CLAUDE_GREEN, alpha=0.08)
    ax.fill(angles, close(ds),  DS_BLACK,  alpha=0.08)

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(categories, fontsize=10, fontweight='medium')
    ax.set_ylim(0, 100)
    ax.set_yticks([20, 40, 60, 80, 100])
    ax.set_yticklabels(['20', '40', '60', '80', '100'], fontsize=7, color='gray')
    ax.yaxis.grid(True, color='#E5E7EB', linewidth=0.6)
    ax.xaxis.grid(True, color='#E5E7EB', linewidth=0.6)
    ax.set_title('Competitive 7-Dimension Radar', pad=28, fontsize=16, fontweight='bold', color='#1F2937')

    ax.legend(loc='upper right', bbox_to_anchor=(1.35, 1.12), frameon=True,
              fancybox=True, edgecolor='#D1D5DB', fontsize=10)

    return save_png(fig, 'fig1_radar.png')


# ============================================================
# FIGURE 2 — Pie Chart: Dilemma Pass Rate
# ============================================================
def fig2_pie():
    sizes = [55, 35, 10]
    colors = [GREEN_PASS, YELLOW_PARTIAL, RED_FAIL]
    labels = ['Passed (55%)', 'Partial (35%)', 'Failed (10%)']
    explode = (0.02, 0.02, 0.08)

    fig, ax = plt.subplots(figsize=(7, 6))
    fig.patch.set_facecolor('white')

    wedges, texts, autotexts = ax.pie(
        sizes, explode=explode, labels=labels, colors=colors,
        autopct='%1.0f%%', startangle=140, pctdistance=0.6,
        wedgeprops={'edgecolor': 'white', 'linewidth': 2.5},
        textprops={'fontsize': 12, 'fontweight': 'medium'}
    )

    for at in autotexts:
        at.set_fontsize(13)
        at.set_fontweight('bold')
        at.set_color('white' if at.get_text() != '55%' else '#1F2937')

    # Center circle for donut effect
    centre_circle = plt.Circle((0, 0), 0.38, fc='white', edgecolor='#E5E7EB', linewidth=1)
    ax.add_artist(centre_circle)
    ax.text(0, 0, 'Dilemma\nPass Rate', ha='center', va='center',
            fontsize=13, fontweight='bold', color='#374151')

    ax.set_title('Dilemma Scenario Pass Rate', pad=20, fontsize=16, fontweight='bold', color='#1F2937')

    # Legend
    legend_labels = ['[PASS] Coach handles dilemma correctly',
                     '[PARTIAL] Coach struggles but recovers',
                     '[FAIL] Coach gives harmful/incorrect advice']
    ax.legend(wedges, legend_labels, loc='lower center', bbox_to_anchor=(0.5, -0.18),
              ncol=1, frameon=False, fontsize=9)

    return save_png(fig, 'fig2_pie.png')


# ============================================================
# FIGURE 3 — Line Chart: Response Time Trend
# ============================================================
def fig3_line():
    weeks = ['W1', 'W2', 'W3', 'W4', 'W5']
    times = [12.0, 8.0, 7.3, 6.0, 15.0]
    target = 8.0

    fig, ax = plt.subplots(figsize=(8, 5))
    fig.patch.set_facecolor('white')
    ax.set_facecolor('#FAFAFA')

    # Target zone
    ax.axhline(y=target, color=GREEN_PASS, linestyle='--', linewidth=2, alpha=0.8, label=f'Target ({target}s)')
    ax.fill_between(range(len(weeks)), 0, target, alpha=0.06, color=GREEN_PASS)

    # Line + markers
    x = range(len(weeks))
    ax.plot(x, times, color=EC_BLUE, linewidth=2.8, marker='o', markersize=10,
            markerfacecolor='white', markeredgewidth=2.5, markeredgecolor=EC_BLUE,
            zorder=5, label='ExpressCoach')

    # Annotate W5 spike
    ax.annotate('!! Crisis\nIntervention',
                xy=(4, 15.0), xytext=(4, 16.2),
                ha='center', fontsize=9, color=RED_FAIL, fontweight='bold',
                arrowprops=dict(arrowstyle='->', color=RED_FAIL, lw=1.5))

    # Fill area under line
    ax.fill_between(x, times, alpha=0.12, color=EC_BLUE)

    ax.set_xticks(x)
    ax.set_xticklabels(weeks, fontsize=12, fontweight='medium')
    ax.set_ylabel('Avg Response Time (seconds)', fontsize=12, color='#374151')
    ax.set_ylim(0, 18)
    ax.yaxis.grid(True, color='#E5E7EB', linewidth=0.6)
    ax.set_title('Response Time Trend (W1–W5)', pad=16, fontsize=16, fontweight='bold', color='#1F2937')

    ax.legend(frameon=True, fancybox=True, edgecolor='#D1D5DB', fontsize=10, loc='upper left')

    # Value labels
    for xi, ti in zip(x, times):
        color = RED_FAIL if ti > target else GREEN_PASS
        ax.text(xi, ti + 0.5, f'{ti}s', ha='center', fontsize=10, fontweight='bold', color=color)

    return save_png(fig, 'fig3_line.png')


# ============================================================
# FIGURE 4 — Bar Chart: Three-Version Selection Rate
# ============================================================
def fig4_bar():
    versions = ['Gentle\n(温和)', 'Firm\n(坚定)', 'High-EQ\n(高情商)']
    rates = [28, 25, 47]
    colors = ['#93C5FD', '#F59E0B', '#34D399']

    fig, ax = plt.subplots(figsize=(8, 5.5))
    fig.patch.set_facecolor('white')
    ax.set_facecolor('#FAFAFA')

    bars = ax.bar(versions, rates, color=colors, edgecolor='white', linewidth=1.5,
                  width=0.55, zorder=3)

    # Value labels
    for bar, rate in zip(bars, rates):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1.2,
                f'{rate}%', ha='center', fontsize=14, fontweight='bold', color='#374151')

    # Add subtle horizontal lines
    ax.yaxis.grid(True, color='#E5E7EB', linewidth=0.6)
    ax.set_axisbelow(True)
    ax.set_ylim(0, 58)
    ax.set_ylabel('Selection Rate (%)', fontsize=12, color='#374151')
    ax.set_title('Coaching Style Preference (n=120 Users)', pad=16, fontsize=16, fontweight='bold', color='#1F2937')

    # Style annotations
    annotations = ['Safety-first,\navoid conflict', 'Direct,\nboundary-setting', 'Empathetic,\nrelationship-building']
    for bar, anno in zip(bars, annotations):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() / 2, anno,
                ha='center', va='center', fontsize=9, color='white', fontweight='bold', alpha=0.95)

    # Legend
    from matplotlib.patches import Patch
    legend_els = [Patch(facecolor=colors[0], label='Gentle — non-confrontational'),
                  Patch(facecolor=colors[1], label='Firm — assertive boundaries'),
                  Patch(facecolor=colors[2], label='High-EQ — empathetic mirroring')]
    ax.legend(handles=legend_els, loc='upper center', bbox_to_anchor=(0.5, -0.10),
              ncol=3, frameon=False, fontsize=9)

    return save_png(fig, 'fig4_bar.png')


# ============================================================
# FIGURE 5 — Heatmap: Strategy Matrix (4 Dimensions × 5 Intents)
# ============================================================
def fig5_heatmap():
    dims = ['Accuracy', 'Empathy', 'Speed', 'Safety']
    intents = ['Inform', 'Coach', 'Support', 'Challenge', 'Reflect']

    # Strategy intensity matrix (0–100)
    data = np.array([
        [90, 75, 50, 30, 40],   # Accuracy
        [40, 85, 95, 55, 80],   # Empathy
        [70, 60, 45, 35, 25],   # Speed
        [85, 70, 80, 65, 60],   # Safety
    ])

    fig, ax = plt.subplots(figsize=(9, 5.5))
    fig.patch.set_facecolor('white')

    cmap = plt.cm.YlOrRd  # Yellow → Orange → Red
    im = ax.imshow(data, cmap=cmap, aspect='auto', vmin=0, vmax=100)

    # Cell labels
    for i in range(len(dims)):
        for j in range(len(intents)):
            val = data[i, j]
            color = 'white' if val > 60 else '#1F2937'
            ax.text(j, i, f'{val}', ha='center', va='center',
                    fontsize=12, fontweight='bold', color=color)

    ax.set_xticks(range(len(intents)))
    ax.set_xticklabels(intents, fontsize=11, fontweight='medium')
    ax.set_yticks(range(len(dims)))
    ax.set_yticklabels(dims, fontsize=11, fontweight='medium')

    # Move x ticks to top
    ax.xaxis.tick_top()
    ax.xaxis.set_label_position('top')

    # Grid
    for i in range(len(dims) + 1):
        ax.axhline(y=i - 0.5, color='white', linewidth=2)
    for j in range(len(intents) + 1):
        ax.axvline(x=j - 0.5, color='white', linewidth=2)

    # Colorbar
    cbar = fig.colorbar(im, ax=ax, shrink=0.82, pad=0.02)
    cbar.set_label('Strategy Intensity', fontsize=11, color='#374151')
    cbar.ax.tick_params(labelsize=9)

    ax.set_title('Strategy Matrix: Dimension × Intent', pad=28, fontsize=16, fontweight='bold', color='#1F2937')

    # Subtitles
    ax.set_xlabel('User Intent →', fontsize=11, color='#6B7280', labelpad=8)
    ax.xaxis.set_label_position('top')

    return save_png(fig, 'fig5_heatmap.png')


# ============================================================
# Main
# ============================================================
if __name__ == '__main__':
    print('Generating W5 figures...\n')

    paths = [
        ('Figure 1 — 7-Dimension Competitive Radar', fig1_radar()),
        ('Figure 2 — Dilemma Pass Rate Pie',         fig2_pie()),
        ('Figure 3 — Response Time Trend (W1–W5)',    fig3_line()),
        ('Figure 4 — Coaching Style Preference',      fig4_bar()),
        ('Figure 5 — Strategy Matrix Heatmap',        fig5_heatmap()),
    ]

    # Embed in Excel
    xlsx_path = os.path.join(OUT, 'W5_Figures.xlsx')
    embed_to_excel([(t, p) for t, p in paths], xlsx_path)

    print('\n[Done] All 5 charts generated and embedded in Excel.')
    print(f'   Output directory: {OUT}')
